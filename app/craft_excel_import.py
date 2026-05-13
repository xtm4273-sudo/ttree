"""
从已下载并人工填写后的「报价单」Excel 中批量学习入库。
约定：列「学习入库」填写 ADD（大小写不敏感）表示将该行写入工艺库；
     工艺标题优先取「人工确认标题」（旧版表头），否则取「匹配工艺（首选）」；
     SFI / 单位优先取「人工确认*」列（旧版），否则取「SFI编码」「单位」；
     若报价单含「解析描述」列则用作工艺说明补充，否则说明字段为空。
"""
from __future__ import annotations

import os
from io import BytesIO
from typing import BinaryIO

from openpyxl import load_workbook

import sys

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from app.craft_library import batch_add_to_library
from app.learn_history import append_learn_event


def _header_map(ws) -> dict[str, int]:
    """首行表头 -> 1-based 列号。"""
    m: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        v = ws.cell(1, col).value
        if v is None:
            continue
        name = str(v).strip()
        if name:
            m[name] = col
    return m


def parse_quotation_excel_for_learning(
    file_or_path: str | bytes | BinaryIO,
) -> tuple[list[dict], list[str]]:
    """
    解析报价单工作簿，返回 (待入库条目列表, 警告/跳过信息)。
    每条 dict: sfi_code, title, description, unit
    """
    if isinstance(file_or_path, str):
        wb = load_workbook(file_or_path, data_only=True, read_only=True)
    elif isinstance(file_or_path, bytes):
        wb = load_workbook(BytesIO(file_or_path), data_only=True, read_only=True)
    else:
        wb = load_workbook(file_or_path, data_only=True, read_only=True)

    try:
        if "报价单" not in wb.sheetnames:
            return [], [f"未找到工作表「报价单」，现有: {wb.sheetnames}"]
        ws = wb["报价单"]
        headers = _header_map(ws)
        if "学习入库" not in headers:
            return [], ["表头缺少「学习入库」，无法导入。请使用系统生成的报价单再填写。"]
        col_action = headers["学习入库"]
        col_title_legacy = headers.get("人工确认标题")
        col_title_new = headers.get("匹配工艺（首选）")
        if not col_title_legacy and not col_title_new:
            return [], ["表头缺少「匹配工艺（首选）」或「人工确认标题」，无法导入。"]

        col_sfi_legacy = headers.get("人工确认SFI")
        col_sfi_enquiry = headers.get("SFI编码")
        col_unit_legacy = headers.get("人工确认单位")
        col_unit_enquiry = headers.get("单位")
        col_desc = headers.get("解析描述")

        entries: list[dict] = []
        notes: list[str] = []

        for r in range(2, ws.max_row + 1):
            action = ws.cell(r, col_action).value
            if action is None or str(action).strip().upper() != "ADD":
                continue
            title = ""
            if col_title_legacy:
                title_cell = ws.cell(r, col_title_legacy).value
                title = str(title_cell).strip() if title_cell else ""
            if not title and col_title_new:
                title_cell = ws.cell(r, col_title_new).value
                title = str(title_cell).strip() if title_cell else ""
            if not title:
                notes.append(f"第{r}行: 学习入库=ADD 但工艺标题列为空，已跳过")
                continue
            sfi = ""
            if col_sfi_legacy:
                v = ws.cell(r, col_sfi_legacy).value
                sfi = str(v).strip() if v is not None else ""
            if not sfi and col_sfi_enquiry:
                v = ws.cell(r, col_sfi_enquiry).value
                sfi = str(v).strip() if v is not None else ""
            unit = "LOT"
            if col_unit_legacy:
                v = ws.cell(r, col_unit_legacy).value
                if v is not None and str(v).strip():
                    unit = str(v).strip()
            elif col_unit_enquiry:
                v = ws.cell(r, col_unit_enquiry).value
                if v is not None and str(v).strip():
                    unit = str(v).strip()
            desc = ""
            if col_desc:
                v = ws.cell(r, col_desc).value
                if v is not None:
                    desc = str(v).strip()[:2000]
            entries.append(
                {
                    "sfi_code": sfi,
                    "title": title,
                    "description": desc,
                    "unit": unit,
                }
            )
        return entries, notes
    finally:
        wb.close()


def import_learning_rows_from_quotation_excel(
    file_or_path: str | bytes | BinaryIO,
    data_dir: str = "data",
) -> tuple[int, list[str]]:
    """
    从已填报价单导入学习条目并写入工艺库。
    返回 (成功写入条数, 全部说明含解析阶段与 batch_add 错误)。
    """
    entries, parse_notes = parse_quotation_excel_for_learning(file_or_path)
    if not entries:
        return 0, parse_notes
    n, batch_errs, added_meta = batch_add_to_library(entries, data_dir=data_dir)
    for meta in added_meta:
        append_learn_event(
            {
                "source": "excel_batch",
                "craft_id": meta.get("craft_id"),
                "craft_title": (meta.get("craft_title") or "")[:500],
                "craft_sfi": meta.get("craft_sfi") or "",
            }
        )
    return n, parse_notes + batch_errs
