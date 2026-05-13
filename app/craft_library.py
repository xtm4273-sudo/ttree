"""
工艺库加载、向量化、检索、增量更新模块
"""
import json
import os
import re
import numpy as np
import faiss
from openpyxl import load_workbook

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config
from app.llm_client import get_embed_client


def craft_entry_dedupe_key(item: dict) -> tuple[str, str]:
    """用于模板行与用户增量行的去重（SFI 大写 + 标题去首尾空格）。"""
    sfi = (item.get("sfi_code") or "").strip().upper()
    title = (item.get("title") or "").strip()
    return (sfi, title)


def ensure_craft_full_text(item: dict) -> None:
    """与 load_craft_library 一致：编码 + 标题 + 可选 detail，用 ' - ' 拼接。"""
    parts = [item.get("sfi_code") or "", item.get("title") or ""]
    detail = (item.get("detail") or "").strip()
    if detail:
        parts.append(detail)
    item["full_text"] = " - ".join(parts).strip() or (item.get("title") or "").strip()


def load_user_added_entries_from_disk(data_dir: str = "data") -> list[dict]:
    """从 craft_library.json 读出 source=user_added 的条目（不含 id，供合并）。"""
    data_path = os.path.join(data_dir, "craft_library.json")
    if not os.path.isfile(data_path):
        return []
    with open(data_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    out: list[dict] = []
    for it in data:
        if it.get("source") != "user_added":
            continue
        title = (it.get("title") or "").strip()
        if not title:
            continue
        out.append(
            {
                "sfi_code": it.get("sfi_code") or "",
                "title": title,
                "detail": it.get("detail") or "",
                "unit": it.get("unit") or "LOT",
                "qty_template": it.get("qty_template"),
                "category": it.get("category") or "",
                "source": "user_added",
            }
        )
    return out


def merge_template_with_saved_user_entries(
    template_items: list[dict], data_dir: str = "data"
) -> list[dict]:
    """
    将模板 Excel 解析结果与磁盘上已保存的 user_added 条目合并（去重后模板在前），
    并重赋 id、统一 full_text。用于「从模板重建索引」时不丢失学习条目。
    """
    seen: set[tuple[str, str]] = {craft_entry_dedupe_key(x) for x in template_items}
    merged: list[dict] = [dict(x) for x in template_items]
    for u in load_user_added_entries_from_disk(data_dir):
        k = craft_entry_dedupe_key(u)
        if k in seen:
            continue
        seen.add(k)
        merged.append(dict(u))
    for i, it in enumerate(merged):
        it["id"] = i
        if it.get("source") != "user_added":
            it["source"] = "template"
        ensure_craft_full_text(it)
    return merged


def load_craft_library(excel_path: str = None) -> list[dict]:
    """
    从万邦报价单模板Excel中提取工艺库条目。

    返回:
    [
        {
            "id": 0,
            "sfi_code": "1.EH.1.1",
            "title": "CRANE SERVICE / SHORE CRANE HOURLY USAGE",
            "detail": "SHORE CRANE SERVICE",
            "unit": "LIFT",
            "qty_template": None,
            "category": "1",
            "full_text": "1.EH.1.1 CRANE SERVICE / SHORE CRANE HOURLY USAGE - SHORE CRANE SERVICE",
        },
        ...
    ]
    """
    path = excel_path or config.CRAFT_LIBRARY_PATH
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb["Quotation"]

    items = []
    current_title_row = None

    for row in ws.iter_rows(min_row=2, values_only=False):
        a_val = str(row[0].value).strip() if row[0].value else ""
        c_val = str(row[2].value).strip() if row[2].value else ""
        d_val = str(row[3].value).strip() if row[3].value else ""
        f_val = row[5].value if row[5].value else None

        if not c_val:
            continue

        # 判断行类型
        is_sfi_row = bool(re.match(r'^\d{1,2}\.[A-Z]{1,4}\.\d', a_val))
        is_category = bool(re.match(r'^\d{1,2}$', a_val))

        if is_category:
            continue

        if is_sfi_row:
            current_title_row = {
                "sfi_code": a_val,
                "title": c_val,
                "detail": "",
                "unit": d_val,
                "qty_template": f_val,
                "category": a_val.split(".")[0],
            }
            items.append(current_title_row)

        elif current_title_row and not is_category:
            # 明细行：补充描述
            if current_title_row["detail"]:
                current_title_row["detail"] += " | " + c_val
            else:
                current_title_row["detail"] = c_val

            if d_val and not current_title_row["unit"]:
                current_title_row["unit"] = d_val
            if f_val and not current_title_row["qty_template"]:
                current_title_row["qty_template"] = f_val

    wb.close()

    # 生成完整文本（用于embedding）和ID
    for i, item in enumerate(items):
        item["id"] = i
        item["source"] = "template"
        ensure_craft_full_text(item)

    return items


def get_embeddings(texts: list[str], batch_size: int = 6) -> np.ndarray:
    """
    调用Embedding API获取文本向量。
    支持OpenAI兼容接口。
    """
    client = get_embed_client()

    all_embeddings = []
    for i in range(0, len(texts), batch_size):
        batch = texts[i:i + batch_size]
        response = client.embeddings.create(
            model=config.EMBED_MODEL,
            input=batch,
        )
        batch_embeddings = [item.embedding for item in response.data]
        all_embeddings.extend(batch_embeddings)

    return np.array(all_embeddings, dtype=np.float32)


def build_vector_index(craft_items: list[dict], save_dir: str = "data"):
    """
    为工艺库构建FAISS向量索引。
    """
    os.makedirs(save_dir, exist_ok=True)

    texts = [item["full_text"] for item in craft_items]

    print(f"[向量化] 正在为 {len(texts)} 条工艺生成embedding...")
    embeddings = get_embeddings(texts)
    print(f"[向量化] 完成，维度: {embeddings.shape}")

    # 归一化（内积 = 余弦相似度）
    faiss.normalize_L2(embeddings)

    # 构建索引
    dim = embeddings.shape[1]
    index = faiss.IndexFlatIP(dim)
    index.add(embeddings)

    # 保存
    index_path = os.path.join(save_dir, "craft_library.index")
    data_path = os.path.join(save_dir, "craft_library.json")

    faiss.write_index(index, index_path)
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(craft_items, f, ensure_ascii=False, indent=2)

    print(f"[向量化] 索引已保存: {index_path}")
    print(f"[向量化] 数据已保存: {data_path}")

    return index


def load_vector_index(data_dir: str = "data"):
    """加载已保存的向量索引和工艺库数据"""
    index_path = os.path.join(data_dir, "craft_library.index")
    data_path = os.path.join(data_dir, "craft_library.json")

    if not os.path.exists(index_path) or not os.path.exists(data_path):
        return None, None

    index = faiss.read_index(index_path)
    with open(data_path, "r", encoding="utf-8") as f:
        craft_items = json.load(f)

    return index, craft_items


def search_similar(query_text: str, index, craft_items: list[dict], top_k: int = None) -> list[dict]:
    """
    向量检索：找到与query最相似的top-K条工艺。
    """
    k = top_k or config.TOP_K

    query_emb = get_embeddings([query_text])
    faiss.normalize_L2(query_emb)

    scores, indices = index.search(query_emb, k)

    results = []
    for score, idx in zip(scores[0], indices[0]):
        if idx < 0:
            continue
        results.append({
            "item": craft_items[idx],
            "score": float(score),
        })

    return results


def batch_search_similar(query_texts: list[str], index, craft_items: list[dict], top_k: int = None) -> list[list[dict]]:
    """
    批量向量检索：一次API调用获取所有embedding，再批量FAISS检索。
    比逐条调用 search_similar 快 N 倍。
    """
    k = top_k or config.TOP_K
    query_embs = get_embeddings(query_texts)
    faiss.normalize_L2(query_embs)

    scores_batch, indices_batch = index.search(query_embs, k)

    all_results = []
    for scores, indices in zip(scores_batch, indices_batch):
        results = []
        for score, idx in zip(scores, indices):
            if idx < 0:
                continue
            results.append({
                "item": craft_items[idx],
                "score": float(score),
            })
        all_results.append(results)

    return all_results


def add_to_library(new_entry: dict, data_dir: str = "data") -> tuple[bool, int | None]:
    """
    增量更新：将人工确认的新条目加入工艺库。
    （学习进化机制的核心）

    Args:
        new_entry: {
            "title": "...",
            "description": "...",
            "unit": "...",
            "sfi_code": "..." 或 None,
        }
        data_dir: 数据目录

    Returns:
        (是否成功, 新条目 craft id)；失败时为 (False, None)。
    """
    index_path = os.path.join(data_dir, "craft_library.index")
    data_path = os.path.join(data_dir, "craft_library.json")

    if not os.path.exists(index_path) or not os.path.exists(data_path):
        return False, None

    title = (new_entry.get("title") or "").strip()
    if not title:
        return False, None

    probe = {"sfi_code": new_entry.get("sfi_code") or "", "title": title}
    # 加载现有数据
    index = faiss.read_index(index_path)
    with open(data_path, "r", encoding="utf-8") as f:
        craft_items = json.load(f)

    nk = craft_entry_dedupe_key(probe)
    for it in craft_items:
        if craft_entry_dedupe_key(it) == nk:
            print(f"[工艺库更新] 跳过重复: SFI={probe['sfi_code']!r} title={title[:40]!r}")
            return False, None

    new_id = len(craft_items)
    new_item = {
        "id": new_id,
        "sfi_code": new_entry.get("sfi_code") or "",
        "title": title,
        "detail": (new_entry.get("description") or "").strip(),
        "unit": (new_entry.get("unit") or "LOT") or "LOT",
        "qty_template": None,
        "category": "",
        "source": "user_added",
    }
    ensure_craft_full_text(new_item)
    full_text = new_item["full_text"]

    # 向量化新条目
    new_emb = get_embeddings([full_text])
    faiss.normalize_L2(new_emb)

    # 加入索引
    index.add(new_emb)
    craft_items.append(new_item)

    # 保存
    faiss.write_index(index, index_path)
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(craft_items, f, ensure_ascii=False, indent=2)

    print(f"[工艺库更新] 新增条目 #{new_id}: {title}")
    return True, new_id


def batch_add_to_library(
    entries: list[dict], data_dir: str = "data"
) -> tuple[int, list[str], list[dict]]:
    """
    批量入库（单次 embedding 批量调用）。每条 entry 同 add_to_library 字段。
    返回 (成功条数, 错误/跳过说明列表, 成功写入条目的摘要列表 craft_id/title/sfi_code)。
    """
    index_path = os.path.join(data_dir, "craft_library.index")
    data_path = os.path.join(data_dir, "craft_library.json")
    errors: list[str] = []
    if not os.path.exists(index_path) or not os.path.exists(data_path):
        return 0, ["索引或 craft_library.json 不存在"], []

    index = faiss.read_index(index_path)
    with open(data_path, "r", encoding="utf-8") as f:
        craft_items = json.load(f)

    existing_keys = {craft_entry_dedupe_key(it) for it in craft_items}
    pending: list[dict] = []
    seen_batch: set[tuple[str, str]] = set()

    for idx, raw in enumerate(entries):
        title = (raw.get("title") or "").strip()
        if not title:
            errors.append(f"第{idx + 1}条: 缺少标题")
            continue
        item = {
            "sfi_code": (raw.get("sfi_code") or "").strip(),
            "title": title,
            "detail": (raw.get("description") or raw.get("detail") or "").strip(),
            "unit": (raw.get("unit") or "LOT") or "LOT",
            "qty_template": None,
            "category": "",
            "source": "user_added",
        }
        k = craft_entry_dedupe_key(item)
        if k in existing_keys or k in seen_batch:
            errors.append(f"第{idx + 1}条: 与库内或本批重复 ({k[0]!r}, {k[1][:30]!r})")
            continue
        seen_batch.add(k)
        existing_keys.add(k)
        pending.append(item)

    if not pending:
        return 0, errors, []

    start_id = len(craft_items)
    for j, it in enumerate(pending):
        it["id"] = start_id + j
        ensure_craft_full_text(it)

    texts = [it["full_text"] for it in pending]
    embs = get_embeddings(texts)
    faiss.normalize_L2(embs)
    index.add(embs)
    craft_items.extend(pending)

    faiss.write_index(index, index_path)
    with open(data_path, "w", encoding="utf-8") as f:
        json.dump(craft_items, f, ensure_ascii=False, indent=2)

    print(f"[工艺库更新] 批量新增 {len(pending)} 条")
    added_meta = [
        {"craft_id": it["id"], "craft_title": it.get("title", ""), "craft_sfi": it.get("sfi_code") or ""}
        for it in pending
    ]
    return len(pending), errors, added_meta
