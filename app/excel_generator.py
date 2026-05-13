"""
Excel报价单生成模块
全量清单模式：每条解析条目必出一行；匹配不确定时保留 Top1 供核对。
"""
import os
import re
import copy
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import sys
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
import config


# 样式
HEADER_FONT = Font(bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

# 与 match_engine._compute_matches 默认 llm_reason 对齐，用于 Excel「匹配原因」展示映射
_INTERNAL_DEFAULT_LLM_REASON = "规则/向量直出"

# 与主表列名一致，供页面预览 / 合并导出
QUOTATION_PREVIEW_COLUMNS = [
    "SFI编码",
    "询价标题",
    "匹配工艺（首选）",
    "单位",
    "数量",
    "置信度%",
    "状态",
    "匹配原因",
    "处理状态",
    "学习入库",
]


def _preview_row_craft_text(result: dict) -> str:
    """与主表「匹配工艺（首选）」列逻辑一致（不含页面覆盖字段）。"""
    best = result.get("best_match")
    is_new = result.get("is_new_item", False)
    suggested = result.get("suggested_entry")
    if is_new and suggested:
        return suggested.get("title", result.get("enquiry_title", "")) or ""
    if best:
        return best.get("craft_title", "") or ""
    return "(未匹配)"


def match_results_to_preview_dataframe(results: list[dict]):
    """生成页面可编辑预览表（与导出主表列一致，不含单价/小计/人工确认列）。"""
    import pandas as pd

    rows: list[dict] = []
    for result in results:
        best = result.get("best_match")
        conf = result.get("confidence", 0)
        is_new = result.get("is_new_item", False)
        review_status = result.get("review_status", "")
        needs_human = result.get("needs_human_review", False)

        ov = (result.get("quotation_craft_override") or "").strip()
        craft_text = ov if ov else _preview_row_craft_text(result)

        if is_new:
            status = "库无候选"
        elif needs_human or review_status == "PENDING_REVIEW":
            status = "待确认"
        elif conf >= config.CONFIDENCE_LEVELS["high"]:
            status = "自动"
        elif conf >= config.CONFIDENCE_LEVELS["medium"]:
            status = "建议复核"
        else:
            status = "需确认"

        proc_status = review_status or ("待审核" if needs_human else "")
        if proc_status == "PENDING_REVIEW":
            proc_status = "待审核"
        elif proc_status == "OK":
            proc_status = "正常"

        rows.append(
            {
                "SFI编码": result.get("enquiry_sfi") or "",
                "询价标题": result.get("enquiry_title", ""),
                "匹配工艺（首选）": craft_text,
                "单位": (
                    (result.get("suggested_entry") or {}).get("unit", "")
                    if is_new and result.get("suggested_entry")
                    else (best.get("unit", "") if best else result.get("unit", ""))
                ),
                "数量": result.get("quantity") if result.get("quantity") is not None else "",
                "置信度%": "—" if is_new else conf,
                "状态": status,
                "匹配原因": format_matching_note_for_business(result),
                "处理状态": proc_status,
                "学习入库": (result.get("quotation_learn_action") or ""),
            }
        )
    return pd.DataFrame(rows, columns=QUOTATION_PREVIEW_COLUMNS)


def _df_cell_str(val) -> str:
    import pandas as pd
    import math

    if val is None or (isinstance(val, float) and math.isnan(val)) or pd.isna(val):
        return ""
    return str(val).strip()


def _df_cell_qty(val):
    import pandas as pd
    import math

    if val is None or val == "" or (isinstance(val, float) and math.isnan(val)) or pd.isna(val):
        return None
    if isinstance(val, (int, float)):
        if isinstance(val, float) and val.is_integer():
            return int(val)
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        q = float(s.replace(",", ""))
        return int(q) if q.is_integer() else q
    except ValueError:
        return None


def merge_preview_dataframe_into_match_results(
    df,
    results: list[dict],
    enquiry_items: list[dict] | None = None,
) -> list[dict]:
    """将页面表格中的编辑写回 match_results（及 enquiry_items）。"""
    if df is None or not len(results):
        return results
    n = min(len(df), len(results))
    merged: list[dict] = []
    for i in range(n):
        r = copy.deepcopy(results[i])
        row = df.iloc[i]
        sfi = _df_cell_str(row.get("SFI编码"))
        title = _df_cell_str(row.get("询价标题"))
        craft = _df_cell_str(row.get("匹配工艺（首选）"))
        unit = _df_cell_str(row.get("单位"))
        qty = _df_cell_qty(row.get("数量"))
        learn = _df_cell_str(row.get("学习入库"))

        r["enquiry_sfi"] = sfi or None
        r["enquiry_title"] = title
        if craft:
            r["quotation_craft_override"] = craft
        else:
            r.pop("quotation_craft_override", None)
        r["unit"] = unit
        r["quantity"] = qty
        r["quotation_learn_action"] = learn

        if enquiry_items and i < len(enquiry_items):
            eq = enquiry_items[i]
            eq["sfi_code"] = sfi or None
            eq["title"] = title or eq.get("title", "")
            eq["unit"] = unit or eq.get("unit")
            eq["quantity"] = qty

        merged.append(r)
    merged.extend(copy.deepcopy(results[n:]))
    return merged


def _strip_legacy_technical_suffixes(text: str) -> str:
    """去掉历史版本写入 llm_reason 的技术性尾缀，便于旧数据导出时仍较可读。"""
    if not text:
        return ""
    s = text.strip()
    s = re.sub(r"\s*\(与第二候选分差[\d.]+分，建议人工确认\)\s*$", "", s)
    s = re.sub(
        r"\s*\|\s*层级风险：[^|]*$",
        "",
        s,
    )
    return s.strip()


def _map_decision_reason_for_display(raw: str) -> str:
    """将门控/决策中的技术说明转为业务可读短句（仅用于 Excel 展示）。"""
    s = (raw or "").strip()
    if not s:
        return ""
    if s == "默认向量直出":
        return "已按工艺库相似度排序，当前推荐为排序第一的首选工艺。"
    if s == "已关闭LLM精排开关":
        return "已关闭深度比对，当前按工艺库相似度自动推荐首选。"
    if s == "无候选":
        return "无可用工艺条目可供比对。"
    if s == "向量检索无候选":
        return "工艺库中未找到相近的标准工艺条目。"
    if "SFI完全一致且向量分" in s and "直接命中" in s:
        return "询价与工艺库中某条工艺的编码（SFI）一致，且描述相近，已优先匹配该条。"
    if "SFI完全一致但向量分" in s and "继续门控判断" in s:
        return "询价与工艺库编码一致，但文字描述相近度一般，系统已继续自动判定。"
    if "Top1向量分" in s and "低于阈值" in s:
        return "系统判断询价与首选工艺的相似度一般，已启用深度比对以挑选更合适条目。"
    if "Top1与Top2分差" in s:
        return "前两名工艺与询价都较接近、区分不明显，已启用深度比对。"
    if "询价SFI与Top1候选SFI冲突" in s:
        return "客户询价编码与系统首选工艺的编码归属不一致，已启用深度比对。"
    if "且分差" in s and "向量直出" in s:
        return "系统判断当前首选明显优于其他条目，已自动采用相似度排序结果。"
    return s


def format_matching_note_for_business(result: dict) -> str:
    """
    生成写入报价单「匹配原因」列的展示文案（面向业务人员，非审计/调试口径）。
    """
    is_new = result.get("is_new_item", False)
    suggested = result.get("suggested_entry")
    best = result.get("best_match")

    if is_new and suggested:
        desc = (suggested.get("description") or "")[:80].strip()
        if desc:
            return f"工艺库暂无对应标准项，以下为系统参考说明：{desc}"
        return "工艺库暂无对应标准项，请按公司主数据补录或手工指定工艺。"

    if is_new and not suggested:
        return (
            _map_decision_reason_for_display(result.get("decision_reason") or "")
            or "工艺库中未找到相近标准工艺，请人工指定或补录工艺库。"
        )

    raw = ""
    if best:
        raw = (best.get("llm_reason") or "").strip()
    else:
        return _map_decision_reason_for_display(result.get("decision_reason") or "") or "无首选工艺，请人工处理。"

    raw = _strip_legacy_technical_suffixes(raw)

    if raw == _INTERNAL_DEFAULT_LLM_REASON:
        return "按工艺库相似度自动推荐当前首选。"

    if raw.startswith("LLM调用失败"):
        return "智能分析暂不可用，当前按相似度排序；请人工核对首选工艺是否准确。"

    mapped = _map_decision_reason_for_display(raw)
    if mapped != raw:
        return mapped

    if len(raw) > 120:
        raw = raw[:117] + "…"

    return raw or "请结合客户描述与首选工艺名称自行核对。"


def generate_quotation_excel(match_results: list[dict], output_path: str) -> str:
    """
    生成报价单Excel（全量行，不丢条目）。
    """
    wb = Workbook()
    _build_quotation_sheet(wb, match_results)
    _build_review_sheet(wb, match_results)
    _build_new_items_sheet(wb, match_results)
    _build_summary_sheet(wb, match_results)
    wb.save(output_path)
    return output_path


def _build_quotation_sheet(wb: Workbook, results: list[dict]):
    """主报价表：全量清单（与页面预览列一致）。"""
    ws = wb.active
    ws.title = "报价单"

    headers = [
        ("SFI编码", 12),
        ("询价标题", 36),
        ("匹配工艺（首选）", 40),
        ("单位", 8),
        ("数量", 6),
        ("置信度%", 7),
        ("状态", 12),
        ("匹配原因", 38),
        ("处理状态", 12),
        ("学习入库", 10),
    ]

    for col, (name, width) in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER
        ws.column_dimensions[get_column_letter(col)].width = width

    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A2"

    row = 2
    for result in results:
        best = result.get("best_match")
        conf = result.get("confidence", 0)
        is_new = result.get("is_new_item", False)
        suggested = result.get("suggested_entry")
        review_status = result.get("review_status", "")
        needs_human = result.get("needs_human_review", False)

        sfi = result.get("enquiry_sfi") or ""
        ws.cell(row, 1, sfi).border = THIN_BORDER

        ws.cell(row, 2, result.get("enquiry_title", "")).border = THIN_BORDER

        ov = (result.get("quotation_craft_override") or "").strip()
        if ov:
            craft_text = ov
        elif is_new and suggested:
            craft_text = suggested.get("title", result.get("enquiry_title", ""))
        elif best:
            craft_text = best.get("craft_title", "")
        else:
            craft_text = "(未匹配)"
        ws.cell(row, 3, craft_text).border = THIN_BORDER

        if is_new and suggested:
            unit = suggested.get("unit", "")
        elif best:
            unit = best.get("unit", "")
        else:
            unit = result.get("unit", "")
        ws.cell(row, 4, unit).border = THIN_BORDER

        qty = result.get("quantity")
        ws.cell(row, 5, qty if qty is not None and qty != "" else "").border = THIN_BORDER

        conf_cell = ws.cell(row, 6, conf if not is_new else "—")
        conf_cell.border = THIN_BORDER
        conf_cell.alignment = Alignment(horizontal="center")
        if conf < 60 and not is_new:
            conf_cell.font = Font(bold=True, color="CC0000")

        if is_new:
            status = "库无候选"
        elif needs_human or review_status == "PENDING_REVIEW":
            status = "待确认"
        elif conf >= config.CONFIDENCE_LEVELS["high"]:
            status = "自动"
        elif conf >= config.CONFIDENCE_LEVELS["medium"]:
            status = "建议复核"
        else:
            status = "需确认"
        ws.cell(row, 7, status).border = THIN_BORDER

        ws.cell(row, 8, format_matching_note_for_business(result)).border = THIN_BORDER

        proc_status = review_status or ("待审核" if needs_human else "")
        if proc_status == "PENDING_REVIEW":
            proc_status = "待审核"
        elif proc_status == "OK":
            proc_status = "正常"
        ws.cell(row, 9, proc_status).border = THIN_BORDER

        learn_val = result.get("quotation_learn_action") or ""
        ws.cell(row, 10, learn_val).border = THIN_BORDER

        for c in range(1, len(headers) + 1):
            ws.cell(row, c).alignment = Alignment(vertical="top", wrap_text=True)

        row += 1

    row += 2
    ws.cell(row, 1, "说明:").font = Font(bold=True, size=9)
    row += 1
    notes = [
        "本表为全量清单：解析出的每条询价均对应一行。",
        "状态为「待确认」且处理状态为待审核时：请在页面上核对「匹配工艺（首选）」等字段后再导出。",
        "库无候选：工艺库中未找到相近标准工艺，请在工艺库补项或手工指定。",
        "学习入库：在页面表格中将「匹配工艺（首选）」等列填写完整后，将「学习入库」列填 ADD，"
        "导出 Excel 并上传，可由系统批量写入工艺库（见侧栏说明）。",
    ]
    for t in notes:
        ws.cell(row, 1, t).font = Font(size=9)
        row += 1


def _build_review_sheet(wb: Workbook, results: list[dict]):
    """需审核条目：待确认或无候选"""
    ws = wb.create_sheet("需审核")

    headers = ["SFI编码", "询价标题", "最佳匹配", "置信度", "原因", "处理状态", "建议操作"]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 36
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 22

    row = 2
    for result in results:
        if not result.get("needs_human_review") and not result.get("is_new_item"):
            continue

        best = result.get("best_match")
        conf = result.get("confidence", 0)
        is_new = result.get("is_new_item", False)

        ws.cell(row, 1, result.get("enquiry_sfi") or "")
        ws.cell(row, 2, result.get("enquiry_title", ""))
        ov = (result.get("quotation_craft_override") or "").strip()
        if ov:
            craft_disp = ov
        elif best:
            craft_disp = best.get("craft_title", "(无)")
        else:
            craft_disp = "(无)"
        ws.cell(row, 3, craft_disp)
        ws.cell(row, 4, conf if not is_new else 0)
        ws.cell(row, 5, format_matching_note_for_business(result))
        rs = result.get("review_status", "") or ""
        if rs == "PENDING_REVIEW":
            rs_disp = "待审核"
        elif rs == "OK":
            rs_disp = "正常"
        else:
            rs_disp = rs
        ws.cell(row, 6, rs_disp)

        if is_new:
            action = "补工艺库或手工填写匹配工艺"
        elif conf < config.CONFIDENCE_LEVELS["low"]:
            action = "匹配可能有误，请手动选择"
        else:
            action = "建议复核匹配是否正确"
        ws.cell(row, 7, action)
        row += 1

    if row == 2:
        ws.cell(2, 1, "（当前无待审核项）")


def _build_new_items_sheet(wb: Workbook, results: list[dict]):
    """工艺库无向量候选（库无候选）"""
    ws = wb.create_sheet("库无候选")

    headers = ["询价标题", "SFI编码", "建议操作"]
    for col, name in enumerate(headers, 1):
        cell = ws.cell(1, col, name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 28

    row = 2
    for result in results:
        if not result.get("is_new_item"):
            continue
        ws.cell(row, 1, result.get("enquiry_title", ""))
        ws.cell(row, 2, result.get("enquiry_sfi") or "")
        ws.cell(row, 3, "补充工艺库后重跑或人工指定")
        row += 1

    if row == 2:
        ws.cell(2, 1, "（无库无候选项）")


def _build_summary_sheet(wb: Workbook, results: list[dict]):
    """统计汇总"""
    ws = wb.create_sheet("统计")

    total = len(results)
    high = sum(1 for r in results if r.get("confidence", 0) >= 80 and not r.get("is_new_item"))
    med = sum(1 for r in results if 60 <= r.get("confidence", 0) < 80 and not r.get("is_new_item"))
    low = sum(1 for r in results if 40 <= r.get("confidence", 0) < 60 and not r.get("is_new_item"))
    no_cand = sum(1 for r in results if r.get("is_new_item"))
    pending = sum(1 for r in results if r.get("review_status") == "PENDING_REVIEW")

    data = [
        ["指标", "数量", "占比"],
        ["总条目数", total, "100%"],
        ["高置信度（>=80）", high, f"{high*100//max(total,1)}%"],
        ["中等（60-79）", med, f"{med*100//max(total,1)}%"],
        ["偏低（40-59）", low, f"{low*100//max(total,1)}%"],
        ["库无候选", no_cand, f"{no_cand*100//max(total,1)}%"],
        ["待人工确认", pending, f"{pending*100//max(total,1)}%"],
        [],
        ["导出完整性", f"{total} 行", "与匹配输入条数一致"],
    ]

    for r, row_data in enumerate(data, 1):
        for c, val in enumerate(row_data, 1):
            cell = ws.cell(r, c, val)
            if r == 1:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 15
